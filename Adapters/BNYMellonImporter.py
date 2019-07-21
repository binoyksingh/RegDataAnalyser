import fnmatch
import os
from datetime import datetime
import csv

from openpyxl import load_workbook

from Modules import RTS27_LEI_Company_Map_Module
from Modules import RTS27_Table_Records_Module
from Modules_DB_Readers import RTS27_Prod_Class_DB_Reader_Module
from Modules_DB_Writers import RTS27_DB_Writer_Module
from Modules_DB_Readers import HistMktData_DB_Reader_Module
from Utilities import RTS27_Utilities

#
# PRIMARY Key to Join Tables 2,4,6 :
# Company Name
# ISIN
# Trade Date
# CCY
# Test - Binoynnn

def getTable3_ObjectFromRow (rowId, wsheet,table2_rec) :
    table3_rec_new = RTS27_Table_Records_Module.RTS27_Table3()
    transaction_price = str(wsheet.cell(row=rowId, column=5).value)
    time_of_execution = str(wsheet.cell(row=rowId, column=6).value)
    transaction_size = str(wsheet.cell(row=rowId, column=7).value)

    table3_rec_new.setSimpleAverageExecutedPrice(str(wsheet.cell(row=rowId, column=3).value))
    table3_rec_new.setTotalValueExecuted(str(wsheet.cell(row=rowId, column=4).value))
    table3_rec_new.setPrice(str(wsheet.cell(row=rowId, column=5).value))
    table3_rec_new.setTimeOfExecution(str(wsheet.cell(row=rowId, column=6).value))
    table3_rec_new.setTransactionSize(str(wsheet.cell(row=rowId, column=7).value))
    table3_rec_new.setTradingSystem(str(wsheet.cell(row=rowId, column=8).value))
    table3_rec_new.setTradingMode(str(wsheet.cell(row=rowId, column=9).value))
    table3_rec_new.setTradingPlatform(str(wsheet.cell(row=rowId, column=10).value))

    #if ( transaction_price != 'None' ):
    #    print "For rowId" + str(rowId) + ", Transaction price : " + transaction_price + "; Time of Execution : " + time_of_execution
    return table3_rec_new

bnymellon_source_path = "/Users/sarthakagarwal/PycharmProjects/MifidDataAnalyser/Source/BNYMellon/Unzipped source"
rtsdb = RTS27_DB_Writer_Module.RTS27_DB_Writer()
histdata_db_reader = HistMktData_DB_Reader_Module.HistMktData_DB_Reader()


bnymellonfilenames = []
for root, dirnames, filenames in os.walk(bnymellon_source_path):
    for filename in fnmatch.filter(filenames, '*.xlsx'):
        bnymellonfilenames.append(os.path.join(root, filename))

#bnymellonfilenames = bnymellonfilenames[0:3]
source_firm_group_name = "BNYMellon"
source_company_name=""
rts_db_rd = RTS27_Prod_Class_DB_Reader_Module.RTS27_Prod_Class_DB_Reader()
transaction_count = 0

table_switches = RTS27_Utilities.RTS27_TableSwitches("N", "N", "Y", "N", "N") #Table 1, Table 2, Table 3, Table 4, Table 6

fileId = 0
list_of_table2_records = []
list_of_table6_records = []
with open('BNYM_RTS27_T3_TxResults.csv', 'w') as writeFile:
    writer = csv.writer(writeFile)
    for filename in bnymellonfilenames:
        #print("Processing file" + os.path.basename(filename))
        wb = load_workbook(filename)
        wsheet = wb.worksheets[0] # reading the first sheet only

        CurrentFinancialInstrumentAtRow = 0
        PrevFinancialInstrumentAtRow = 0
        formatted_date = ""
        company_code = ""

        for rowcount in range(1, wsheet.max_row + 1):
            if ((wsheet.cell(row=rowcount , column=2)).value == "Type of Execution Venue"):
                # Building Table 1 object
                table1_rec = RTS27_Table_Records_Module.RTS27_Table1()
                table1_rec.setSourceCompanyCode(str(str(wsheet.cell(row=rowcount+1 , column=2).value).decode('ascii', errors='ignore')))
                company_code = str(wsheet.cell(row=rowcount+1 , column=2).value)
                lei_comp_obj = RTS27_LEI_Company_Map_Module.RTS27_LEI_Company_Map()
                source_company_name = lei_comp_obj.getMap()[company_code]
                table1_rec.setSourceCompanyGroupName(source_firm_group_name)
                table1_rec.setSourceCompanyName(source_company_name)
                rawdate = datetime.strptime((wsheet.cell(row=5 , column=2)).value, '%Y-%m-%d')
                formatted_date = datetime.strftime(rawdate, "%Y-%m-%d")
                table1_rec.setTradeDate(formatted_date)
                table1_rec.setCountryOfCompetentAuthority((wsheet.cell(row=3 , column=2)).value)
                table1_rec.setMarketSegmentID((wsheet.cell(row=4, column=2)).value)
                table1_rec.setFileName(os.path.basename(filename))
                table1_rec.setFailedTransactionsNumber((wsheet.cell(row=8, column=2)).value)

            if ((wsheet.cell(row=rowcount , column=2)).value == "Type of Financial Instrument"):

                CurrentFinancialInstrumentAtRow = rowcount
                # Found an instrument, now to build the objects
                # Building Table 2 object
                table2_rec = RTS27_Table_Records_Module.RTS27_Table2(rts_db_rd.getCfi_assetclass_map(), rts_db_rd.getCfi_char_map())
                table2_rec.setTradeDate(formatted_date)
                table2_rec.setInstrumentName(str(str(wsheet.cell(row=rowcount+1 , column=2).value).decode('ascii', errors='ignore')))
                table2_rec.setISIN(str(wsheet.cell(row=rowcount+2 , column=2).value))
                table2_rec.setInstrumentClassification(str(wsheet.cell(row=rowcount+4 , column=2).value))
                table2_rec.setCurrency(str(wsheet.cell(row=rowcount+5 , column=2).value))

                table2_rec.setSourceCompanyName(source_company_name)
                table2_rec.setFileName(os.path.basename(filename))
                table2_rec.setTradeDate(formatted_date)

                table2_rec.setFileId(company_code +"_"+ formatted_date +"_"+ table2_rec.ISIN + "_" + table2_rec.CURRENCY)

                # -----------------------------------------------------------
                # Building Table 4
                table4_rec_new = RTS27_Table_Records_Module.RTS27_Table4()
                table4_rec_new.SOURCE_COMPANY_NAME = source_company_name
                table4_rec_new.FILENAME = os.path.basename(filename)
                table4_rec_new.TRADE_DATE = formatted_date
                table4_rec_new.FILE_ID = table2_rec.FILE_ID
                table4_rec_new.INSTRUMENT_NAME = table2_rec.INSTRUMENT_NAME
                table4_rec_new.ISIN = table2_rec.ISIN
                table4_rec_new.CURRENCY = table2_rec.CURRENCY

                if (wsheet.cell(row=rowcount + 24, column=2).value is not None):
                    table4_rec_new.SIMPLE_AVERAGE_TRANSACTION_PRICE = str(wsheet.cell(row=rowcount + 24, column=2).value)
                if (wsheet.cell(row=rowcount + 25, column=2).value is not None):
                    table4_rec_new.VOLUME_WEIGHTED_TRANSACTION_PRICE = str(wsheet.cell(row=rowcount + 25, column=2).value)
                if (wsheet.cell(row=rowcount + 26, column=2).value is not None):
                    table4_rec_new.HIGHEST_EXECUTED_PRICE = str(wsheet.cell(row=rowcount + 26, column=2).value)
                if (wsheet.cell(row=rowcount + 27, column=2).value is not None):
                    table4_rec_new.LOWEST_EXECUTED_PRICE = str(wsheet.cell(row=rowcount + 27, column=2).value)

                # -----------------------------------------------------------
                # Building Table 6
                table6_rec_new = RTS27_Table_Records_Module.RTS27_Table6()
                table6_rec_new.SOURCE_COMPANY_NAME = source_company_name
                table6_rec_new.FILENAME = os.path.basename(filename)
                table6_rec_new.TRADE_DATE = formatted_date
                table6_rec_new.FILE_ID = table2_rec.FILE_ID
                table6_rec_new.INSTRUMENT_NAME = table2_rec.INSTRUMENT_NAME
                table6_rec_new.ISIN = table2_rec.ISIN
                table6_rec_new.CURRENCY = table2_rec.CURRENCY

                if (wsheet.cell(row=rowcount+30 , column=2).value is not None ) :
                    table6_rec_new.NUMBER_OF_ORDER_OR_REQUEST_FOR_QUOTE = int(wsheet.cell(row=rowcount+30 , column=2).value)
                if (wsheet.cell(row=rowcount+31 , column=2).value is not None ) :
                    table6_rec_new.NUMBER_OF_TRANSACTIONS_EXECUTED = int(wsheet.cell(row=rowcount+31 , column=2).value)
                if (wsheet.cell(row=rowcount+32 , column=2).value is not None ) :
                    table6_rec_new.TOTAL_VALUE_OF_TRANSACTIONS_EXECUTED = str(wsheet.cell(row=rowcount+32 , column=2).value)

                if (wsheet.cell(row=rowcount+33 , column=2).value is not None ) :
                    table6_rec_new.NUMBER_OF_ORDERS_OR_REQUEST_CANCELLED_OR_WITHDRAWN = int(wsheet.cell(row=rowcount+33 , column=2).value)

                if (wsheet.cell(row=rowcount+34 , column=2).value is not None ) :
                    table6_rec_new.NUMBER_OF_ORDERS_OR_REQUEST_MODIFIED = (wsheet.cell(row=rowcount+34 , column=2).value)

                if (wsheet.cell(row=rowcount+35 , column=2).value is not None ) :
                    table6_rec_new.MEDIAN_TRANSACTION_SIZE =  str(wsheet.cell(row=rowcount+35 , column=2).value)

                if (wsheet.cell(row=rowcount+36 , column=2).value is not None ) :
                    table6_rec_new.MEDIAN_SIZE_OF_ALL_ORDERS_OR_REQUESTS_FOR_QUOTE = str(wsheet.cell(row=rowcount+36 , column=2).value)

                if (wsheet.cell(row=rowcount+37 , column=2).value is not None ) :
                    table6_rec_new.NUMBER_OF_DESIGNATED_MARKET_MAKER = str(wsheet.cell(row=rowcount+37 , column=2).value)

                table1_rec.FILE_ID = table2_rec.FILE_ID

                # -----------------------------------------------------------
                # Building Table 3
                table3_records = []
                # print "Tenor, Transaction Size, BNYM Executed Price, Mid Market Price, Margin"

                for rowId in range(rowcount + 10, rowcount + 21):
                    table3_rec = getTable3_ObjectFromRow(rowId, wsheet, table2_rec)
                    if (table3_rec is not None and table3_rec.PRICE!= 'None' and table3_rec.PRICE!= 0.0):
                        table3_rec.setISIN(table2_rec.ISIN)
                        table3_rec.setSourceCompanyName(source_company_name)
                        table3_rec.setSourceCompanyName(source_company_name)
                        table3_rec.setFileName(os.path.basename(filename))
                        table3_rec.setTradeDate(formatted_date)
                        table3_rec.setFileId(table2_rec.FILE_ID)
                        table3_rec.setInstrumentName(table2_rec.INSTRUMENT_NAME)
                        table3_rec.setCurrency(table3_rec.CURRENCY)
                        table3_rec.setCurrencyPair(table3_rec.CURRENCY)
                        # print (table3_rec.getAttrArrayTable())

                        value_date = datetime.strptime(table2_rec.VALUE_DATE, '%Y-%m-%d')
                        trade_date = datetime.strptime(table2_rec.TRADE_DATE, '%Y-%m-%d')
                        tenor_days = value_date - trade_date

                        fx_mid_price_for_ccy_tenor = histdata_db_reader.getFXFwdHistMktDataForCcy(table2_rec.CURRENCY,
                                           table3_rec.getTimeOfExecutionInESTWithoutDayLightSavings(), tenor_days)

                        if (fx_mid_price_for_ccy_tenor!=0) :

                            #print str(tenor_days.days) + "," + str(table3_rec.TRANSACTION_SIZE) + "," + \
                            #      str(table3_rec.PRICE) + "," + str(fx_mid_price_for_ccy_tenor) + "," + \
                            #      str(abs(float(table3_rec.PRICE) - fx_mid_price_for_ccy_tenor))

                            ccy_pair = (table2_rec.CCY_PAIR).strip()

                            #if (min(table2_rec.CCY1.strip(), table2_rec.CCY2.strip()) == table3_rec.CURRENCY.strip()):
                            #    table3_rec.setConvertedPrice(table3_rec.PRICE)
                            #else:
                            #    table3_rec.setConvertedPrice(1 / float(table3_rec.PRICE))

                            table3_rec.setMidMarketRate(fx_mid_price_for_ccy_tenor)

                            table3_rec.setCurrencyPair(table2_rec.CCY_PAIR)
                            table3_rec.setTenorDays(tenor_days.days)

                            abs_price_diff = abs(float(table3_rec.PRICE) - fx_mid_price_for_ccy_tenor)
                            table3_rec.setAbsPriceDiff(str(format(abs_price_diff, '.18f')))
                            table3_rec.setMarkupAmount(abs_price_diff * float(table3_rec.TRANSACTION_SIZE))
                            table3_rec.setTCAPerformed()

                            if (table2_rec.CCY1.strip() != "USD"):
                                usd_ccy_pair = table2_rec.CCY1.strip() + "USD"
                                fx_spot_for_ccy = histdata_db_reader.getFXSpotHistMktDataForCcyPair(usd_ccy_pair,
                                                                                                    table3_rec.getTimeOfExecutionInESTWithoutDayLightSavings(),
                                                                                                    )
                                markup_usd = float(fx_spot_for_ccy) * float(table3_rec.MARKUP_AMOUNT)

                                table3_rec.setMarkUpUSD(markup_usd)
                            else:
                                table3_rec.setMarkUpUSD(table3_rec.MARKUP_AMOUNT)

                            print (table3_rec.getAttrArrayTable())
                            #writer.writerow(table3_rec.getAttrArrayTable())

                            table3_records.append(table3_rec)
                # -----------------------------------------------------------

                if (table_switches.PROCESS_TABLE_1 == "Y"):
                    # Writing to Table 1 to Database
                    rtsdb.Write_to_Table1(table1_rec)

                if (table_switches.PROCESS_TABLE_2 == "Y") :
                    # Writing to Table 2 to Database
                    rtsdb.Write_to_Table2(table2_rec)

                if (table_switches.PROCESS_TABLE_3 == "Y") :
                    # Writing to Table 3 to Database
                    for table3_rec_temp in table3_records:
                        rtsdb.Write_to_Table3(table3_rec_temp)

                if (table_switches.PROCESS_TABLE_4 == "Y") :
                    # Writing to Table 4 to Database
                    rtsdb.Write_to_Table4(table4_rec_new)

                if (table_switches.PROCESS_TABLE_6 == "Y") :
                    # Writing to Table 6 to Database
                    rtsdb.Write_to_Table6(table6_rec_new)