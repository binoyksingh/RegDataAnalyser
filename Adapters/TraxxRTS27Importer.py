#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Mon Sep  3 22:29:12 2018

@author: lojinilogesparan
"""

import csv
import os
from datetime import datetime
from Modules import RTS27_DB_Writer_Module, RTS27_Table_Records_Module, RTS27_Utilities
from Modules import RTS27_Prod_Class_DB_Reader_Module
from Modules import RTS27_LEI_Company_Map_Module
from openpyxl import load_workbook
import collections

import fnmatch

def get_immediate_subdirectories(a_dir):
    return [name for name in os.listdir(a_dir)
            if os.path.isdir(os.path.join(a_dir, name))]

rtsdb = RTS27_DB_Writer_Module.RTS27_DB_Writer()
rts_db_rd = RTS27_Prod_Class_DB_Reader_Module.RTS27_Prod_Class_DB_Reader()
lei_comp_obj = RTS27_LEI_Company_Map_Module.RTS27_LEI_Company_Map()

all_file_ids = []
all_file_ids.append("test")
all_file_ids.append("test")

#path = "/Users/lojinilogesparan/Documents/mifid_data/State Street Bank Intl/BESTEX_RTS27_ZMHGNT7ZPKZ3UFZ8EO46_2018Q1/"
path= "/Users/sarthakagarwal/PycharmProjects/MifidDataAnalyser/Source/TRAX"
source_firm_name = "State Street Bank Intl"
dateformat =  '%Y%m%d'

firm_source_folders = get_immediate_subdirectories(path)
#firm_source_folders = firm_source_folders[0:2]

for firm_folder in firm_source_folders:
    print "processing company " + firm_folder

    firm_rts27_folders = get_immediate_subdirectories (path + "/" + firm_folder)

    # First build the InstrumentMap
    # Looping through all the folders, looking for Table 4, and then building the InstrumentMap
    instrumentMap = {}
    for foldername in firm_rts27_folders:

        full_folder_path = path + "/" + firm_folder + "/" + foldername
        if ("TABLE4" in foldername):
            trax_firm_table4_filenames = []
            for root, dirnames, filenames in os.walk(full_folder_path):
                for filename in fnmatch.filter(filenames, '*.CSV'):
                    trax_firm_table4_filenames.append(os.path.join(root, filename))

            for table4_filename in trax_firm_table4_filenames:
                    table4_base_file_name = os.path.basename(table4_filename)
                    trade_date = (table4_base_file_name.split("_")[5]).replace(".CSV", "")
                    with open(table4_filename) as csvfile:
                        readCSV = csv.reader(csvfile, delimiter=',')
                        for row in readCSV:
                            instrumentMap[row[0] + "_" + trade_date] = [row[1]]

            print "Instrument Map is " + str(instrumentMap)

    # Now building Objects Table 2 and Table 6
    for foldername in firm_rts27_folders:

        full_folder_path = path + "/" + firm_folder + "/" + foldername
        if ("TABLE6" in foldername):
            trax_firm_table6_filenames = []
            for root, dirnames, filenames in os.walk(full_folder_path):
                for filename in fnmatch.filter(filenames, '*.CSV'):
                    trax_firm_table6_filenames.append(os.path.join(root, filename))

            for table6_filename in trax_firm_table6_filenames:
                base_file_name = os.path.basename(table6_filename)
                source_firm_lei = (base_file_name.split("_")[4]).replace(".CSV","")
                trade_date = (base_file_name.split("_")[5]).replace(".CSV","")
                source_company_name = lei_comp_obj.getMap()[source_firm_lei]

                if (source_company_name == None) :
                    raise ValueError('Company Name not found in LEI Map')

                rawdate = datetime.strptime(trade_date, '%Y%m%d')
                formatted_date = datetime.strftime(rawdate, "%Y-%m-%d")

                with open(table6_filename, 'rb') as csvfile:
                    data = csv.reader(csvfile, delimiter=',', quotechar='|')
                    rowCount = 0
                    for row in data:
                            rowCount = rowCount + 1
                            if rowCount > 1:

                                # We need to read Table 2 first and get the curreny pair from there
                                # ------------------------------
                                # Building Table 2
                                isin = row[0]
                                isin_key = isin.strip() + "_" + trade_date
                                isin_details = instrumentMap.get(isin_key)
                                isin_ccy = ""
                                if (isin_details != None):
                                    print "found currency in Map"
                                    isin_ccy = isin_details[0]

                                file_id_string = source_firm_lei + "_" + trade_date + "_" + isin + "_" + isin_ccy

                                table2_rec = RTS27_Table_Records_Module.RTS27_Table2(rts_db_rd.getCfi_assetclass_map(), rts_db_rd.getCfi_char_map())
                                table2_rec.setFileId(file_id_string)
                                table2_rec.setISIN(str(isin))
                                table2_rec.setCurrency(isin_ccy)
                                table2_rec.setSourceCompanyName(source_company_name)
                                table2_rec.setFileName(os.path.basename(table6_filename))
                                table2_rec.setTradeDate(formatted_date)
                                table2_rec.setValueDate(None) # Setting as default in-case it doesnt get called later

                                # -----------------------------------------------------------
                                # Building Table 6
                                table6_rec_new = RTS27_Table_Records_Module.RTS27_Table6()
                                table6_rec_new.SOURCE_COMPANY_NAME = source_company_name
                                table6_rec_new.FILENAME = os.path.basename(table6_filename)

                                table6_rec_new.TRADE_DATE = formatted_date

                                table6_rec_new.FILE_ID = file_id_string
                                #table6_rec_new.INSTRUMENT_NAME = ''
                                table6_rec_new.ISIN = isin
                                table6_rec_new.CURRENCY = isin_ccy

                                if (row[1] != ' ' and row[1] is not None):
                                    table6_rec_new.setNumberOfOrderOrRequestForQuote(int(row[1]))

                                if (row[2] != ' ' and table6_rec_new.is_number(row[2]) ):
                                    table6_rec_new.setNumberOfTransactionsExecuted(int(row[2]))

                                if (row[3] != ' ' and table6_rec_new.is_number(row[3])):
                                    table6_rec_new.setTotalValueOfTransactionsExecuted(str(row[3]))

                                if (row[4] != ' ' and table6_rec_new.is_number(row[4])):
                                    table6_rec_new.setNumberOfOrdersOrRequestCancelledOrWithdrawn(int(row[4]))

                                if (row[5] != ' ' and table6_rec_new.is_number(row[5])):
                                    table6_rec_new.setNumberOfOrdersOrRequestModified(row[5])

                                if (row[6] != ' ' and table6_rec_new.is_number(row[6])):
                                    table6_rec_new.setMedianTransactionSize(str(row[6]))

                                if (row[7] != ' ' and table6_rec_new.is_number(row[7])):
                                    table6_rec_new.setMedianSizeOfAllOrdersOrRequestsForQuote(str(row[7]))

                                if (row[8] != ' ' and table6_rec_new.is_number(row[8])):
                                    table6_rec_new.setNumberOfDesignatedMarketMaker(str(row[8]))

                                if ((table6_rec_new.key_fields_blank() == False) and (file_id_string not in all_file_ids)):
                                    #print table6_rec_new.getAttrArray()
                                    # Writing to Table 6 to Database
                                    rtsdb.Write_to_Table6(table6_rec_new)

                                    # print table2_rec.getAttrArray()
                                    # Writing to Table 2
                                    rtsdb.Write_to_Table2(table2_rec)

                                    all_file_ids.append(file_id_string)

            trax_firm_table6_filenames_xslx = []
            for root, dirnames, filenames in os.walk(full_folder_path):
                for filename in fnmatch.filter(filenames, '*.xlsx'):
                    trax_firm_table6_filenames_xslx.append(os.path.join(root, filename))

                for table6_filename in trax_firm_table6_filenames_xslx:
                    source_firm_lei = firm_folder.split("_")[2]
                    source_company_name = lei_comp_obj.getMap()[source_firm_lei]

                    if (source_company_name == None):
                        raise ValueError('Company Name not found in LEI Map')

                    wb = load_workbook(table6_filename)
                    wsheet = wb.worksheets[0]  # reading the first sheet only
                    trade_date = ""

                    for rowcount in range(1, wsheet.max_row + 1):

                        first_col_val = (str)((wsheet.cell(row=rowcount, column=1)).value)

                        if ((first_col_val != '') and (((first_col_val)[0]).isalpha() == False) and (first_col_val!='None') and (first_col_val!='#N/A')):

                            rawdate = datetime.strptime(first_col_val, '%Y-%m-%d %H:%M:%S')
                            trade_date = datetime.strftime(rawdate, "%Y-%m-%d")

                        if (((first_col_val[0]).isalpha() == True) and (first_col_val!='None') and (first_col_val!='Financial Instrument (ISIN)') and (first_col_val!='#N/A') ) :

                            # ------------------------------
                            # Building Table 2
                            isin = (wsheet.cell(row=rowcount , column=1)).value
                            isin_details = instrumentMap.get(isin)
                            isin_ccy = ""
                            if (isin_details != None):
                                isin_ccy = isin_details[0]

                            file_id_string = source_firm_lei + "_" + trade_date + "_" + isin + "_" + isin_ccy

                            table2_rec = RTS27_Table_Records_Module.RTS27_Table2(rts_db_rd.getCfi_assetclass_map(), rts_db_rd.getCfi_char_map())
                            table2_rec.setFileId(file_id_string)
                            table2_rec.setISIN(str(isin))
                            table2_rec.setCurrency(isin_ccy)
                            table2_rec.setSourceCompanyName(source_company_name)
                            table2_rec.setFileName(os.path.basename(table6_filename))
                            table2_rec.setTradeDate(trade_date)
                            table2_rec.setValueDate(None) # Setting as default in-case it doesnt get called later

                            # -----------------------------------------------------------
                            # Building Table 6
                            table6_rec_new = RTS27_Table_Records_Module.RTS27_Table6()
                            table6_rec_new.SOURCE_COMPANY_NAME = source_company_name
                            table6_rec_new.FILENAME = os.path.basename(table6_filename)

                            table6_rec_new.TRADE_DATE = trade_date

                            table6_rec_new.FILE_ID = file_id_string
                            table6_rec_new.ISIN = isin
                            table6_rec_new.CURRENCY = isin_ccy

                            if (((wsheet.cell(row=rowcount , column=2)).value != ' ') and ((wsheet.cell(row=rowcount , column=2)).value is not None)):
                                table6_rec_new.setNumberOfOrderOrRequestForQuote(int((wsheet.cell(row=rowcount , column=2)).value))

                            if (((wsheet.cell(row=rowcount , column=3)).value != ' ') and ((wsheet.cell(row=rowcount , column=3)).value  is not None)):
                                table6_rec_new.setNumberOfTransactionsExecuted(int((wsheet.cell(row=rowcount , column=3)).value))

                            if (((wsheet.cell(row=rowcount , column=4)).value != ' ') and ((wsheet.cell(row=rowcount , column=4)).value is not None)):
                                table6_rec_new.setNumberOfOrdersOrRequestCancelledOrWithdrawn(int((wsheet.cell(row=rowcount , column=4)).value))

                            if (((wsheet.cell(row=rowcount , column=5)).value != ' ') and ((wsheet.cell(row=rowcount , column=5)).value is not None)):
                                table6_rec_new.setNumberOfOrdersOrRequestModified((wsheet.cell(row=rowcount , column=5)).value)

                            if (((wsheet.cell(row=rowcount , column=6)).value != ' ') and ((wsheet.cell(row=rowcount , column=6)).value is not None)):
                                table6_rec_new.setMedianTransactionSize(str((wsheet.cell(row=rowcount , column=6)).value))

                            if (((wsheet.cell(row=rowcount , column=7)).value != ' ') and ((wsheet.cell(row=rowcount , column=7)).value is not None)):
                                table6_rec_new.setMedianSizeOfAllOrdersOrRequestsForQuote(str((wsheet.cell(row=rowcount , column=7)).value))

                            if (((wsheet.cell(row=rowcount , column=8)).value != ' ') and ((wsheet.cell(row=rowcount , column=8)).value is not None)):
                                table6_rec_new.setNumberOfDesignatedMarketMaker(str((wsheet.cell(row=rowcount , column=8)).value))

                            if ((table6_rec_new.key_fields_blank() == False) and (file_id_string not in all_file_ids)):
                                #print table6_rec_new.getAttrArray()
                                # Writing to Table 6 to Database
                                rtsdb.Write_to_Table6(table6_rec_new)

                                # print table2_rec.getAttrArray()
                                # Writing to Table 2
                                rtsdb.Write_to_Table2(table2_rec)

                                all_file_ids.append(file_id_string)

                            rowCount += 1

        if ("TABLE4" in foldername):

            trax_firm_table4_filenames = []
            for root, dirnames, filenames in os.walk(full_folder_path):
                for filename in fnmatch.filter(filenames, '*.CSV'):
                    trax_firm_table4_filenames.append(os.path.join(root, filename))

            for table4_filename in trax_firm_table4_filenames :

                base_file_name = os.path.basename(table4_filename)
                source_firm_lei = base_file_name.split("_")[4]
                trade_date = (base_file_name.split("_")[5]).replace(".CSV","")
                source_company_name = lei_comp_obj.getMap()[source_firm_lei]

                if (source_company_name == None):
                    raise ValueError('Company Name not found in LEI Map')

                rawdate = datetime.strptime(trade_date, '%Y%m%d')
                formatted_date = datetime.strftime(rawdate, "%Y-%m-%d")

                with open(table4_filename, 'rb') as csvfile:
                    data = csv.reader(csvfile, delimiter=',', quotechar='|')
                    rowCount = 0
                    for row in data:
                            #print('Row=' + str(rowCount))
                            rowCount += 1
                            isin = row[0]
                            isin_details = instrumentMap.get(isin)
                            isin_ccy = ""
                            if (isin_details != None):
                                isin_ccy = isin_details[0]

                            file_id_string = source_firm_lei + "_" + trade_date + "_" + isin + "_" + isin_ccy

                            # ------------------------------
                            # Building Table 4
                            table4_rec_new = RTS27_Table_Records_Module.RTS27_Table4()
                            table4_rec_new.SOURCE_COMPANY_NAME = source_company_name
                            table4_rec_new.FILENAME = base_file_name
                            table4_rec_new.FILE_ID = file_id_string
                            table4_rec_new.ISIN = isin
                            table4_rec_new.CURRENCY = isin_ccy

                            if (row[2] != ' '):
                                table4_rec_new.setSimpleAverageTransactionPrice(str(row[2]))
                            if (row[3] != ' '):
                                table4_rec_new.setVolumeWeightedTransactionPrice(str(row[3]))
                            if (row[4] != ' '):
                                table4_rec_new.setHighestExecutedPrice(str(row[4]))
                            if (row[5] != ' '):
                                table4_rec_new.setLowestExecutedPrice(str(row[5]))

                            table4_rec_new.TRADE_DATE = formatted_date

                            #print table4_rec_new.getAttrArray()
                            # Writing to Table 4
                            rtsdb.Write_to_Table4(table4_rec_new)


print [item for item, count in collections.Counter(all_file_ids).items() if count > 1]
print len(all_file_ids)